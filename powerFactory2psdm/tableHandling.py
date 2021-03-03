'''
@author: Kittl
'''
import csv
from openpyxl import Workbook   # Has to be installed manually
from shutil import copyfile   # Copying of files
from datetime import datetime
from os import remove
from os.path import exists
import sqlalchemy as sa

def writeToCsv(dpf, data, filePath, worksheets, exportProfile):
    # csv.register_dialect('kittl', delimiter=',', lineterminator = ';')

    cntWs = 0
    for ws in worksheets[int(exportProfile)-1]:   # Iterate through worksheets
        fileName = filePath + "_" + ws + ".csv"
        if exists(fileName):
            dpf.PrintInfo('The file "%s" already exists. Back it up and overwrite it!' % fileName)
            copyfile(fileName, filePath + "_" + ws + "_bak" + datetime.now().strftime('%Y%m%d-%H%M%S') + ".csv")
            remove(fileName)

        with open(fileName, 'w', newline='') as f:
            w = csv.writer(f, dialect='excel')
            w.writerows(data[cntWs])
            f.close
        cntWs += 1

    # csv.unregister_dialect('kittl')
    return

def writeToXlsx(dpf, data, filePath, worksheets, exportProfile):
    fileName = filePath + ".xlsx"
    if exists(fileName):
        dpf.PrintInfo('The file "%s" already exists. Back it up and overwrite it!' % fileName)
        copyfile(fileName, filePath + "_bak" + datetime.now().strftime('%Y%m%d-%H%M%S') + ".xlsx")
        remove(fileName)

    # ====== Prepare the Workbook ======
    wb = Workbook()
    ws = list()
    for cntWs in range(0, len(worksheets[int(exportProfile)-1])):
        # Create all the worksheets
        if cntWs is 0:
            ws.append(wb.active)
            ws[cntWs].title = worksheets[int(exportProfile)-1][cntWs]
        else:
            ws.append(wb.create_sheet(title=worksheets[int(exportProfile)-1][cntWs]))

    # ====== Push the data to the worksheets ======
    if len(data) is not len(ws):
        dpf.PrintError("The length of the data list and the number of worksheets do not match!")
        return 1

    for cntWs in range(0, len(data)):  # Iterate through worksheets
        for cntObj in range(0, len(data[cntWs])):  # Iterate through rows
            for cntCol in range(0, len(data[cntWs][cntObj])):  # Iterate through columns
                ws[cntWs].cell(row=cntObj+1, column=cntCol+1).value = data[cntWs][cntObj][cntCol]

    # ====== Save the Workbook ======
    wb.save(fileName)
    return 0

def writeToDb(dpf, data, dbEngine, projectName, exportProfile, tables):
    # List the names of existing tables
    listOfTables = list()
    for table in dbEngine._metadata.tables.values():
        listOfTables.append(table.name)

    for cntWs in range(0, len(tables[int(exportProfile)-1])):   # Iterate through tables
        tableName = projectName + "_" + tables[int(exportProfile)-1][cntWs]
        if tableName in listOfTables:
            # ------ Table already exists ------
            dpf.PrintInfo('Table "%s" already exists, back it up and overwrite it!' % tableName)
            #  http://www.paulsprogrammingnotes.com/2014/01/clonecopy-table-schema-from-one.html
            srcTable = sa.Table(tableName, dbEngine._metadata)
            destTable = sa.Table(tableName + "_bak" + datetime.now().strftime('%Y%m%d-%H%M%S'),
                                         dbEngine._metadata)
            for col in srcTable.columns:
                destTable.append_column(col.copy())
            destTable.create()
            srcTable.delete()

    return